# =========================================================
# Excel导出
# export_questions.py
#
# 功能：
# 1. 接收网页传入的 questions
# 2. 根据Excel模板创建Excel
# 3. 从第8行开始写入
# 4. 复制模板第7行样式
# 5. 自动换行
# 6. 保存到 excel/目录
#
# 本文件只负责Excel导出
# 不负责AI出题
# 不负责法规知识库
# 不负责生成题目
# =========================================================

import re
from pathlib import Path
from copy import copy

from openpyxl import load_workbook


# =========================================================
# 1. 项目根目录
# =========================================================

BASE_DIR = Path(__file__).resolve().parent


# =========================================================
# 2. Excel输出目录
# =========================================================

EXCEL_DIR = BASE_DIR / "excel"


# =========================================================
# 3. Excel模板
# =========================================================

TEMPLATE_CANDIDATES = [
    BASE_DIR / "导入模版 (1).xlsx",
    BASE_DIR / "导入模版 (2).xlsx",
    BASE_DIR / "导入模版.xlsx",
]


# =========================================================
# 4. Excel配置
# =========================================================

# 从第8行开始写
EXCEL_START_ROW = 8

# 模板第7行作为样式
STYLE_SOURCE_ROW = 7

# A-K共11列
EXCEL_COLUMN_COUNT = 11


# =========================================================
# 5. 创建Excel目录
# =========================================================

EXCEL_DIR.mkdir(
    parents=True,
    exist_ok=True
)


# =========================================================
# 6. 安全文件名
# =========================================================

def safe_filename(name):
    """
    Windows文件名非法字符替换。
    """

    return re.sub(
        r'[\\/:*?"<>|]',
        "_",
        str(name).strip()
    )


# =========================================================
# 7. 查找Excel模板
# =========================================================

def find_template():
    """
    查找Excel导入模板。
    """

    for template in TEMPLATE_CANDIDATES:

        if template.exists():

            print()
            print("✅ 找到Excel模板：")
            print(template.resolve())

            return template.resolve()

    print()
    print("❌ 找不到Excel导入模板！")
    print()

    print("程序尝试寻找：")

    for template in TEMPLATE_CANDIDATES:
        print(
            f"    {template}"
        )

    print()

    print(
        "请把Excel模板放到 "
        "export_questions.py 所在目录。"
    )

    return None


# =========================================================
# 8. 获取Excel输出文件
# =========================================================

def get_excel_file(
    group_name="煤矿",
    category_name="法规"
):
    """
    生成Excel文件名。

    例如：

    AI生成题目.xlsx
    """

    safe_group = safe_filename(
        group_name
    )

    safe_category = safe_filename(
        category_name
    )

    filename = (
        f"{safe_group}+"
        f"{safe_category}.xlsx"
    )

    return (
        EXCEL_DIR / filename
    ).resolve()


# =========================================================
# 9. 复制单元格样式
# =========================================================

def copy_cell_style(
    source_cell,
    target_cell
):
    """
    复制模板单元格样式。
    """

    if source_cell.has_style:

        target_cell._style = copy(
            source_cell._style
        )

    if source_cell.number_format:

        target_cell.number_format = (
            source_cell.number_format
        )

    if source_cell.alignment:

        target_cell.alignment = copy(
            source_cell.alignment
        )

    if source_cell.protection:

        target_cell.protection = copy(
            source_cell.protection
        )


# =========================================================
# 10. 复制模板行样式
# =========================================================

def copy_template_row_style(
    ws,
    source_row,
    target_row
):
    """
    将模板第7行的样式复制到目标行。
    """

    for column in range(
        1,
        EXCEL_COLUMN_COUNT + 1
    ):

        source = ws.cell(
            row=source_row,
            column=column
        )

        target = ws.cell(
            row=target_row,
            column=column
        )

        copy_cell_style(
            source,
            target
        )


# =========================================================
# 11. 创建Excel
# =========================================================

def create_excel_from_template(
    questions,
    output_file,
    template_file
):
    """
    根据题目列表创建Excel。
    """

    print()
    print(
        "===================================="
    )

    print(
        "正在生成Excel..."
    )

    print(
        f"模板：{template_file}"
    )

    print(
        f"输出：{output_file}"
    )

    print(
        f"题目数量：{len(questions)}"
    )

    print(
        "===================================="
    )

    # =====================================================
    # 检查模板
    # =====================================================

    if not template_file.exists():

        print(
            f"❌ 模板不存在：{template_file}"
        )

        return {
            "success": False,
            "message": "Excel模板不存在"
        }

    # =====================================================
    # 打开Excel模板
    # =====================================================

    try:

        wb = load_workbook(
            template_file
        )

    except Exception as e:

        print(
            f"❌ Excel模板读取失败：{e}"
        )

        return {
            "success": False,
            "message": f"Excel模板读取失败：{e}"
        }

    # =====================================================
    # 第一个工作表
    # =====================================================

    ws = wb.worksheets[0]

    # =====================================================
    # 删除第8行以后原来的数据
    # =====================================================

    if ws.max_row >= EXCEL_START_ROW:

        delete_count = (
            ws.max_row
            - EXCEL_START_ROW
            + 1
        )

        ws.delete_rows(
            EXCEL_START_ROW,
            delete_count
        )

    # =====================================================
    # 开始写题目
    # =====================================================

    current_row = EXCEL_START_ROW

    for index, question in enumerate(
        questions,
        start=1
    ):

        # =================================================
        # 防止异常数据
        # =================================================

        if not isinstance(
            question,
            dict
        ):
            print(
                f"⚠️ 第{index}题不是对象，跳过"
            )
            continue

        # =================================================
        # 复制模板第7行样式
        # =================================================

        copy_template_row_style(
            ws,
            STYLE_SOURCE_ROW,
            current_row
        )

        # =================================================
        # A-K
        # =================================================

        values = [

            # A
            index,

            # B
            question.get(
                "title_category_name",
                ""
            ),

            # C
            question.get(
                "subjects",
                ""
            ),

            # D
            question.get(
                "plan_a",
                ""
            ),

            # E
            question.get(
                "plan_b",
                ""
            ),

            # F
            question.get(
                "plan_c",
                ""
            ),

            # G
            question.get(
                "plan_d",
                ""
            ),

            # H
            question.get(
                "plan_e",
                ""
            ),

            # I
            question.get(
                "plan_f",
                ""
            ),

            # J
            question.get(
                "analysis",
                ""
            ),

            # K
            question.get(
                "answer",
                ""
            )
        ]

        # =================================================
        # 写入A-K
        # =================================================

        for column, value in enumerate(
            values,
            start=1
        ):

            cell = ws.cell(
                row=current_row,
                column=column
            )

            cell.value = value

            # 自动换行
            cell.alignment = copy(
                cell.alignment
            )

            cell.alignment = (
                cell.alignment.copy(
                    wrap_text=True
                )
            )

        # =================================================
        # 设置行高
        # =================================================

        ws.row_dimensions[
            current_row
        ].height = 60

        current_row += 1

    # =====================================================
    # 保存Excel
    # =====================================================

    output_file.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    try:

        wb.save(
            output_file
        )

    except PermissionError:

        print()
        print(
            "❌ Excel保存失败！"
        )

        print(
            "请先关闭已经打开的Excel文件，"
        )

        print(
            "然后重新运行。"
        )

        return {
            "success": False,
            "message": "Excel文件正在被占用，请关闭Excel后重试"
        }

    except Exception as e:

        print()
        print(
            f"❌ Excel保存失败：{e}"
        )

        return {
            "success": False,
            "message": f"Excel保存失败：{e}"
        }

    # =====================================================
    # 检查文件
    # =====================================================

    if not output_file.exists():

        print(
            "❌ Excel保存后没有找到文件！"
        )

        return {
            "success": False,
            "message": "Excel保存后文件不存在"
        }

    real_path = output_file.resolve()

    print()
    print(
        "===================================="
    )

    print(
        "✅ Excel生成成功！"
    )

    print()

    print(
        "📄 Excel实际保存位置："
    )

    print(
        real_path
    )

    print(
        "===================================="
    )

    return {
        "success": True,
        "message": "Excel生成成功",
        "excel_file": str(real_path),
        "count": len(questions)
    }


# =========================================================
# 12. main
# =========================================================

def main(
    data=None,
    questions=None,
    group_name="煤矿",
    category_name="法规"
):
    """
    Excel导出统一入口。

    支持：

    方式1：

        main(data={
            "questions": [...]
        })

    方式2：

        main(
            questions=[...]
        )
    """

    print()
    print(
        "===================================="
    )

    print(
        "        Excel题库导出程序"
    )

    print(
        "===================================="
    )

    # =====================================================
    # 处理data
    # =====================================================

    if data is not None:

        if not isinstance(
            data,
            dict
        ):

            return {
                "success": False,
                "message": "Excel导出参数必须是对象"
            }

        # -----------------------------------------------
        # 从data获取questions
        # -----------------------------------------------

        questions = data.get(
            "questions",
            questions
        )

        # -----------------------------------------------
        # 从data获取分组
        # -----------------------------------------------

        group_name = data.get(
            "group_name",
            group_name
        )

        # -----------------------------------------------
        # 从data获取分类
        # -----------------------------------------------

        category_name = data.get(
            "category_name",
            category_name
        )

    # =====================================================
    # 检查questions
    # =====================================================

    if questions is None:

        return {
            "success": False,
            "message": "没有传入questions"
        }

    if not isinstance(
        questions,
        list
    ):

        return {
            "success": False,
            "message": "questions必须是数组"
        }

    print()

    print(
        f"收到题目数量：{len(questions)}"
    )

    # =====================================================
    # 空题库
    # =====================================================

    if not questions:

        print()
        print(
            "⚠️ 没有题目，不生成Excel。"
        )

        return {
            "success": False,
            "message": "没有题目可导出",
            "count": 0
        }

    # =====================================================
    # 查找模板
    # =====================================================

    template_file = find_template()

    if template_file is None:

        return {
            "success": False,
            "message": "找不到Excel导入模板",
            "count": len(questions)
        }

    # =====================================================
    # 获取输出文件
    # =====================================================

    excel_file = get_excel_file(
        group_name=group_name,
        category_name=category_name
    )

    print()

    print(
        "Excel输出文件："
    )

    print(
        excel_file
    )

    # =====================================================
    # 创建Excel
    # =====================================================

    result = create_excel_from_template(
        questions=questions,
        output_file=excel_file,
        template_file=template_file
    )

    # =====================================================
    # 返回
    # =====================================================

    return result


# =========================================================
# 13. 直接运行测试
# =========================================================

if __name__ == "__main__":

    print(
        "请通过main.py调用Excel导出。"
    )